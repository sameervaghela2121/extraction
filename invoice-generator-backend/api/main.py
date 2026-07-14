import asyncio
import hmac
import io
import json
import logging
import os
import re
import time
import uuid
from datetime import datetime, timezone
from pathlib import Path

import cv2
import numpy as np
from dotenv import load_dotenv
from fastapi import Depends, FastAPI, File, Form, Header, HTTPException, UploadFile
from pydantic import BaseModel
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse, StreamingResponse
from google import genai
from google.genai import types
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

load_dotenv()

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("invoice")

# How many PDFs to send to Gemini at once. Higher = faster bulk, but risks rate limits.
MAX_CONCURRENCY = int(os.environ.get("MAX_CONCURRENCY", "5"))
# How many times to retry rate-limited files (each retry halves concurrency + backs off).
RATE_LIMIT_RETRIES = int(os.environ.get("RATE_LIMIT_RETRIES", "4"))
# Gemini 2.5 Flash pricing, USD per 1M tokens (thinking tokens bill as output).
# Override via env if Google changes the price list.
GEMINI_INPUT_PER_1M = float(os.environ.get("GEMINI_INPUT_PER_1M", "0.30"))
GEMINI_OUTPUT_PER_1M = float(os.environ.get("GEMINI_OUTPUT_PER_1M", "2.50"))
USD_TO_INR = float(os.environ.get("USD_TO_INR", "88"))
# ponytail: single hardcoded user (env-overridable); session tokens are in-memory
# and die on restart — same lifetime as the JOBS store. Move to real auth if
# multiple users ever matter.
APP_USER = os.environ.get("APP_USER", "darshilc@thirdrocktechkno.com")
APP_PASSWORD = os.environ.get("APP_PASSWORD", "admin@123")
TOKENS: set[str] = set()
# Gemini call timeout: 30 minutes (ms) so very large documents aren't cut off by the SDK default.
GEMINI_TIMEOUT_MS = int(os.environ.get("GEMINI_TIMEOUT_MS", str(30 * 60 * 1000)))
# MongoDB persistence (optional): every extracted invoice is saved here when the URI is set.
MONGODB_URI = os.environ.get("MONGODB_URI", "")
MONGO_DB = os.environ.get("MONGO_DB", "Data")
MONGO_COLLECTION = os.environ.get("MONGO_COLLECTION", "Invoice")
MONGO_FILES_COLLECTION = os.environ.get("MONGO_FILES_COLLECTION", "Files")
# Laplacian-variance floor for camera photos. Below this the text can't be transcribed
# reliably and Gemini will invent plausible digits — better to ask for a retake.
BLUR_THRESHOLD = float(os.environ.get("BLUR_THRESHOLD", "60"))
# One Gemini response is capped at 65536 output tokens. A PDF holding many invoices
# produces more JSON than that and gets truncated mid-string, so split big PDFs into
# batches of whole invoices and extract each separately.
SPLIT_PAGES = int(os.environ.get("SPLIT_PAGES", "8"))              # only split beyond this
MAX_INVOICES_PER_CALL = int(os.environ.get("MAX_INVOICES_PER_CALL", "5"))
# Where the uploaded PDFs/photos are kept so the library can show them again later.
# Must be a persistent volume in production, or files vanish on redeploy.
STORAGE_DIR = Path(os.environ.get("STORAGE_DIR", "storage")).resolve()
# New uploads go to this private GCS bucket instead of local disk (files uploaded before
# this change stay on STORAGE_DIR and keep working via the local-disk path below). The
# bucket has no public access — /files/{fid}/raw streams bytes through this service, so
# viewing still requires a valid portal login, same as before.
GCS_BUCKET = os.environ.get("GCS_BUCKET", "sameerv-docflow-invoices")
GCS_PREFIX = "gcs://"
# Only these are ever accepted for upload — anything else (text/html, image/svg+xml, etc.)
# is stored as inert bytes so a malicious upload can never be served back as active content.
ALLOWED_UPLOAD_MIMES = {"application/pdf", "image/jpeg", "image/png", "image/webp"}

_mongo_client = None
_gcs_bucket_handle = None


def _gcs_bucket():
    """Lazy singleton GCS bucket handle (picks up ambient/service-account credentials)."""
    global _gcs_bucket_handle
    if _gcs_bucket_handle is None:
        from google.cloud import storage
        _gcs_bucket_handle = storage.Client().bucket(GCS_BUCKET)
    return _gcs_bucket_handle


def _mongo_collection(name: str = None):
    """Lazy singleton client; returns a collection, or None when MONGODB_URI is unset."""
    global _mongo_client
    if not MONGODB_URI:
        return None
    if _mongo_client is None:
        from pymongo import MongoClient
        _mongo_client = MongoClient(MONGODB_URI, serverSelectionTimeoutMS=10_000)
    return _mongo_client[MONGO_DB][name or MONGO_COLLECTION]


def store_file(jid: str, idx: int, filename: str, data: bytes, mime: str = "application/octet-stream") -> str:
    """Upload to the private GCS bucket under <job>/<idx>_<name> and return a "gcs://<object>"
    reference (never a public URL — the bucket has no public access; see get_file_raw).

    The idx prefix keeps two uploads with the same name from colliding, and the
    name is stripped of anything that could climb out of the directory. The stored
    content-type is forced to a safe value unless it's one we actually expect, so a
    malicious upload can never be served back as active content (e.g. text/html).
    """
    safe = re.sub(r"[^A-Za-z0-9._-]", "_", Path(filename or "file").name) or "file"
    object_name = f"{jid}/{idx}_{safe}"
    safe_mime = mime if mime in ALLOWED_UPLOAD_MIMES else "application/octet-stream"
    blob = _gcs_bucket().blob(object_name)
    blob.upload_from_string(data, content_type=safe_mime)
    return f"{GCS_PREFIX}{object_name}"


def invoice_title(invoices: list[dict]) -> str:
    """Name a document after the invoice(s) inside it — '795/26-27 — Olympic Decor LLP'.

    A phone hands us IMG_20260713_154102.jpg, which tells the user nothing. What they
    look for is the invoice number and who issued it.
    """
    first = invoices[0]
    no = str(first.get("invoice_no") or "").strip()
    seller = str(first.get("seller_name") or "").strip()
    # no "+N more" suffix: the row already shows the invoice count as its own pill, and the
    # suffix was what pushed the actual name off a narrow screen
    return " — ".join(p for p in (no, seller) if p)[:120]


def create_file_docs(jid: str, stored: list[dict]) -> list[dict]:
    """Insert one Files doc per upload, up front, marked 'processing'.

    Written at upload time (not at job end) so a user who closes the tab and comes
    back mid-job still sees their files in the library — and so a server restart can
    find the interrupted work and resume it. Fills each `stored` entry with its _id.
    """
    try:
        coll = _mongo_collection(MONGO_FILES_COLLECTION)
        if coll is None or not stored:
            return stored
        now = datetime.now(timezone.utc)
        docs = [{"job_id": jid, "filename": s["filename"], "mime": s["mime"], "size": s["size"],
                 "path": s["path"], "idx": s["idx"], "status": "processing", "error": None,
                 "invoice_count": 0, "renamed": False,
                 # a placeholder until extraction can name it properly
                 "title": s.get("title") or s["filename"],
                 "created_at": now} for s in stored]
        for s, oid in zip(stored, coll.insert_many(docs).inserted_ids):
            s["_id"] = oid
    except Exception:
        log.exception("Job %s: could not pre-register files in MongoDB", jid)
    return stored


def set_file_status(file_id, status: str):
    """Mirror a file's live progress into Mongo so the library reflects it."""
    try:
        coll = _mongo_collection(MONGO_FILES_COLLECTION)
        if coll is not None and file_id is not None:
            coll.update_one({"_id": file_id}, {"$set": {"status": status}})
    except Exception:
        log.exception("could not update status of file %s", file_id)


def save_invoices_to_db(jid: str, invoices: list[dict], errors: list[dict],
                        cost: dict | None = None, stored: list[dict] | None = None):
    """Persist extracted invoices (and failures), and finalise the Files docs that
    create_file_docs() registered at upload time. Never raises — a Mongo outage must
    not fail the extraction job."""
    try:
        coll = _mongo_collection()
        if coll is None:
            log.warning("MONGODB_URI not set; skipping DB save for job %s", jid)
            return
        now = datetime.now(timezone.utc)

        # Everything joins on the upload's index, never on its name.
        errors_by_idx = {e.get("file_idx"): e.get("error") for e in errors}
        file_ids = {}  # file_idx -> ObjectId
        if stored:
            files_coll = _mongo_collection(MONGO_FILES_COLLECTION)
            for s in stored:
                if s.get("_id") is None:
                    continue
                idx = s["idx"]
                file_ids[idx] = s["_id"]
                mine = [i for i in invoices if i.get("file_idx") == idx]
                update = {
                    "status": "failed" if idx in errors_by_idx else "done",
                    "error": errors_by_idx.get(idx),
                    "invoice_count": len(mine),
                }
                # name the document after what's actually on it, unless the user renamed it
                doc = files_coll.find_one({"_id": s["_id"]}, {"renamed": 1})
                if mine and not (doc or {}).get("renamed"):
                    title = invoice_title(mine)
                    if title:
                        update["title"] = title
                files_coll.update_one({"_id": s["_id"]}, {"$set": update})

        # job_cost_* fields are the WHOLE job's Gemini cost, stamped on each of its
        # documents for easy querying — don't sum them across invoices of one job.
        cost = cost or {}
        docs = [{**inv, "job_id": jid, "file_id": file_ids.get(inv.get("file_idx")),
                 "validation": _validate_invoice(inv), "created_at": now, **cost}
                for inv in invoices]
        docs += [{"job_id": jid, "file_id": file_ids.get(e.get("file_idx")),
                  "source_file": e.get("source_file"), "error": e.get("error"),
                  "created_at": now, **cost} for e in errors]
        if docs:
            coll.insert_many(docs)
        log.info("Job %s: saved %d invoice(s) + %d error(s) + %d file(s) to %s",
                 jid, len(invoices), len(errors), len(stored or []), MONGO_DB)
    except Exception:
        log.exception("Job %s: MongoDB save failed (extraction result unaffected)", jid)

app = FastAPI(title="Invoice to Excel")

# CORS: wide open — any origin, method, and header allowed.
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

EXTRACTION_PROMPT = """You are extracting structured data from a document that contains ONE OR MORE
purchase/sales invoices (native-text PDF or scanned image). Read every field visible on each invoice.
Return ONLY minified JSON — a single line with no indentation or extra whitespace (large invoices
truncate otherwise). No markdown fences, no commentary. If the document contains exactly one invoice,
return one JSON object; if it contains multiple invoices, return a JSON array with one object per
invoice. Each object has this shape:

{
  "page": number,
  "invoice_no": string,
  "invoice_date": string,
  "seller_name": string,
  "seller_gstin": string,
  "buyer_name": string,
  "buyer_gstin": string,
  "items": [
    {"description": string, "hsn": string, "qty": number, "unit": string, "rate": number, "amount": number}
  ],
  "taxable_value": number,
  "cgst_rate": string, "cgst_amount": number,
  "sgst_rate": string, "sgst_amount": number,
  "igst_rate": string, "igst_amount": number,
  "round_off": number,
  "grand_total": number,
  "grand_total_words": string,
  "other_fields": {"any additional field label on the document": "its value"}
}

"page" is the 1-based page of THIS PDF on which the invoice starts (the first page of the invoice, not
the page a continued invoice ends on). For a single-page image, "page" is 1. Count pages of the file you
were given, starting at 1 — do not use any page number printed on the document itself.

Use "" or 0 for fields that are genuinely absent from the document. Put anything on the document that
doesn't fit the fixed fields above (IRN, Ack No, vehicle no, eway bill no, bank details, royalty/cess
fields, remarks, etc.) into "other_fields" as label/value pairs, using the label as printed on the document.
Numbers must be plain JSON numbers, not strings, and not formatted with commas.

ACCURACY RULES — follow these exactly:
- Transcribe ONLY what is actually printed on the document. Copy every digit, GSTIN, date and
  amount character-for-character as shown.
- Do NOT calculate, infer, correct, round, or reformat any value. If the printed total looks wrong,
  still copy exactly what is printed — never compute your own.
- If a field is missing, blank, illegible, or you are not certain what it says, use "" (or 0 for a
  number). NEVER guess a value and never fill a field with a plausible-looking placeholder.
- Do not invent fields, line items, or values that are not on the document.
- A GSTIN is exactly 15 characters. Copy it verbatim; if you cannot read all 15 clearly, use "".
- Treat each invoice as fully independent. NEVER copy, merge, or carry over any value (items,
  totals, GSTINs, dates) from one invoice into another. Every page belongs to exactly one invoice —
  a multi-page invoice is still ONE object; a new invoice starts only where a new invoice number /
  header appears.
"""


def preprocess_photo(data: bytes) -> bytes:
    """Auto-crop and deskew a phone photo of a document (classic scanner-app cleanup):
    find the largest 4-corner contour and perspective-warp it flat. On any failure
    (no document found, decode error) the original bytes come back unchanged —
    preprocessing must never block extraction."""
    try:
        img = cv2.imdecode(np.frombuffer(data, np.uint8), cv2.IMREAD_COLOR)
        if img is None:
            return data
        h, w = img.shape[:2]
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        blur = cv2.GaussianBlur(gray, (5, 5), 0)
        edges = cv2.Canny(blur, 50, 150)
        edges = cv2.dilate(edges, np.ones((3, 3), np.uint8))
        contours, _ = cv2.findContours(edges, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

        quad = None
        for c in sorted(contours, key=cv2.contourArea, reverse=True)[:5]:
            peri = cv2.arcLength(c, True)
            approx = cv2.approxPolyDP(c, 0.02 * peri, True)
            # must be a quadrilateral covering a meaningful part of the frame
            if len(approx) == 4 and cv2.contourArea(approx) > 0.2 * h * w:
                quad = approx.reshape(4, 2).astype("float32")
                break
        if quad is None:
            return data

        # order corners: top-left, top-right, bottom-right, bottom-left
        s = quad.sum(axis=1)
        d = np.diff(quad, axis=1).ravel()
        tl, br = quad[np.argmin(s)], quad[np.argmax(s)]
        tr, bl = quad[np.argmin(d)], quad[np.argmax(d)]
        out_w = int(max(np.linalg.norm(br - bl), np.linalg.norm(tr - tl)))
        out_h = int(max(np.linalg.norm(tr - br), np.linalg.norm(tl - bl)))
        if out_w < 50 or out_h < 50:
            return data
        src = np.array([tl, tr, br, bl], dtype="float32")
        dst = np.array([[0, 0], [out_w - 1, 0], [out_w - 1, out_h - 1], [0, out_h - 1]], dtype="float32")
        warped = cv2.warpPerspective(img, cv2.getPerspectiveTransform(src, dst), (out_w, out_h))

        ok, buf = cv2.imencode(".jpg", warped, [cv2.IMWRITE_JPEG_QUALITY, 90])
        return buf.tobytes() if ok else data
    except Exception:
        log.exception("preprocess_photo failed; using original image")
        return data


def photo_sharpness(data: bytes) -> float:
    """Variance of the Laplacian — the standard blur metric. A sharp document scan
    scores in the hundreds; a soft/out-of-focus one collapses toward zero."""
    try:
        img = cv2.imdecode(np.frombuffer(data, np.uint8), cv2.IMREAD_GRAYSCALE)
        if img is None:
            return float("inf")   # can't judge it — don't block it
        return float(cv2.Laplacian(img, cv2.CV_64F).var())
    except Exception:
        log.exception("sharpness check failed; letting the photo through")
        return float("inf")


def _fix_mojibake(value):
    # ponytail: Gemini occasionally double-encodes a character on noisy scanned
    # regions (stamps/QR/handwriting), e.g. "ò" -> "Ã²". Round-tripping through
    # latin-1/utf-8 only succeeds (and only changes anything) when that happened.
    if isinstance(value, str):
        try:
            return value.encode("latin-1").decode("utf-8")
        except (UnicodeDecodeError, UnicodeEncodeError):
            return value
    if isinstance(value, list):
        return [_fix_mojibake(v) for v in value]
    if isinstance(value, dict):
        return {k: _fix_mojibake(v) for k, v in value.items()}
    return value


def _num(v):
    try:
        return float(str(v).replace(",", "").strip())
    except (TypeError, ValueError, AttributeError):
        return None


def _cell(v):
    # openpyxl only writes scalars; flatten lists/dicts Gemini sometimes returns
    # (e.g. multi-line terms, nested bank details) into a readable string.
    if isinstance(v, (list, tuple)):
        return " ; ".join(_cell(x) for x in v)
    if isinstance(v, dict):
        return " ; ".join(f"{k}: {_cell(val)}" for k, val in v.items())
    return v


# Money that sits on the invoice but has no column in our fixed schema — it lands in
# other_fields. Ignoring it made the totals "fail" on perfectly-transcribed invoices.
_ADDS_TO_TOTAL = ("freight", "transport", "packing", "forwarding", "insurance",
                  "cess", "tcs", "loading", "courier", "delivery", "round")
_SUBTRACTS_FROM_TOTAL = ("discount", "rebate", "less", "tds")


def _extra_money(inv: dict) -> float:
    """Net of the charges/deductions Gemini parked in other_fields."""
    total = 0.0
    for label, value in (inv.get("other_fields") or {}).items():
        n = _num(value)
        if n is None:
            continue
        low = str(label).lower()
        if any(w in low for w in _SUBTRACTS_FROM_TOTAL):
            total -= abs(n)
        elif any(w in low for w in _ADDS_TO_TOTAL):
            total += n
    return total


def _validate_invoice(inv: dict) -> str:
    """Cross-check the money math. Returns "OK" when the figures reconcile, else a
    short note. A note does NOT mean the extraction is wrong — the document itself may
    carry charges we can't model, or genuinely not add up."""
    notes = []
    taxable = _num(inv.get("taxable_value"))
    grand = _num(inv.get("grand_total"))

    # 1) sum of line-item amounts should equal the taxable value
    amounts = [_num(it.get("amount")) for it in (inv.get("items") or [])]
    amounts = [a for a in amounts if a is not None]
    if amounts and taxable is not None:
        s = sum(amounts)
        if abs(s - taxable) > max(1.0, abs(taxable) * 0.001):
            notes.append(f"line items total {s:.2f}, taxable value reads {taxable:.2f}")

    # 2) taxable + taxes + round-off + any freight/discount/cess = grand total
    if taxable is not None and grand is not None:
        computed = taxable
        for k in ("cgst_amount", "sgst_amount", "igst_amount", "round_off"):
            n = _num(inv.get(k))
            if n is not None:
                computed += n
        computed += _extra_money(inv)
        if abs(computed - grand) > max(1.0, abs(grand) * 0.001):
            notes.append(f"figures add up to {computed:.2f}, grand total reads {grand:.2f}")

    # 3) GSTINs are exactly 15 chars when present
    for key in ("seller_gstin", "buyer_gstin"):
        g = inv.get(key)
        if g and len(str(g).strip()) != 15:
            notes.append(f"{key} is not 15 characters")

    return "; ".join(notes) if notes else "OK"


VERIFY_SUFFIX = """

IMPORTANT — this document was already read once and the figures did not reconcile:
{issues}
Read it again from scratch. Look very carefully at the digits in the totals, the tax
amounts and the line-item amounts, and copy them EXACTLY as printed. Do not compute or
correct anything: if the document itself does not add up, still transcribe what is
printed. Include every charge shown (freight, discount, cess, TCS, round-off) — put any
that has no field in the schema into other_fields with its printed label."""


def build_custom_fields_prompt(fields: list[dict]) -> str:
    """Render an extra instruction block for admin-defined custom fields (from the
    portal's Extraction Settings page), so they land in other_fields under a stable,
    known key instead of whatever label Gemini would otherwise pick on its own."""
    if not fields:
        return ""
    lines = "\n".join(
        f'- key "{f["key"]}" ({f["label"]}): {f["description"]}' if f.get("description")
        else f'- key "{f["key"]}" ({f["label"]})'
        for f in fields
    )
    return f"""

In addition to the fixed fields above, also specifically look for these additional fields on the
document. Include each one in "other_fields" using EXACTLY the key given below (not the label), and
use "" if it is genuinely not present on the document — do not guess:
{lines}"""


BOUNDARY_PROMPT = """This PDF contains one or more invoices, one after another. An invoice may
span several pages.

Return ONLY a JSON array of integers: the 1-based page numbers on which a NEW invoice BEGINS.
A page begins a new invoice when it shows a new invoice header / invoice number — a continuation
page of the same invoice (more line items, totals, terms, annexure) is NOT a new invoice.
Page 1 always begins an invoice, so the first element is always 1.

Example for a 5-page file holding 3 invoices where the second one runs over two pages:
[1,2,4]

No commentary, no other keys — just the array."""


def pdf_page_count(data: bytes) -> int:
    """Pages in a PDF; 0 when it can't be read (caller then treats it as unsplittable)."""
    try:
        from pypdf import PdfReader
        return len(PdfReader(io.BytesIO(data)).pages)
    except Exception:
        log.exception("could not read PDF page count")
        return 0


def find_invoice_pages(client: genai.Client, filename: str, data: bytes) -> tuple[list[int], dict]:
    """Ask Gemini which pages start a new invoice.

    The whole point is that the answer is a handful of integers, so it cannot truncate —
    unlike a full extraction of a 19-invoice document, which does. Knowing the real
    boundaries is what lets us split without ever cutting an invoice in half.
    """
    usage = {"input": 0, "output": 0}
    log.info("  boundary pass: %s (%.0f KB)", filename, len(data) / 1024)
    response = client.models.generate_content(
        model="gemini-2.5-flash",
        contents=[types.Part.from_bytes(data=data, mime_type="application/pdf"), BOUNDARY_PROMPT],
        config=types.GenerateContentConfig(
            response_mime_type="application/json",
            temperature=0,
            max_output_tokens=8192,   # a list of page numbers — never large
        ),
    )
    u = response.usage_metadata
    if u:
        usage["input"] += u.prompt_token_count or 0
        usage["output"] += (u.candidates_token_count or 0) + (u.thoughts_token_count or 0)

    pages = json.loads(response.text)
    if not isinstance(pages, list):
        raise ValueError(f"boundary pass returned {type(pages).__name__}, expected a list")
    total = pdf_page_count(data)
    # keep only sane, ascending, in-range page numbers; page 1 always starts an invoice
    clean = sorted({int(p) for p in pages if isinstance(p, (int, float)) and 1 <= int(p) <= total})
    if not clean or clean[0] != 1:
        clean = [1] + clean
    log.info("  boundary pass: %d invoice(s) across %d page(s) — starts at %s",
             len(clean), total, clean)
    return clean, usage


def split_pdf_by_invoice(data: bytes, starts: list[int], max_invoices: int) -> list[tuple[bytes, int]]:
    """Cut the PDF into sub-PDFs of at most `max_invoices` whole invoices.

    Splits on invoice boundaries, never mid-invoice — a 2-page invoice stays intact.
    Returns [(pdf_bytes, first_page)] in document order.
    """
    from pypdf import PdfReader, PdfWriter

    reader = PdfReader(io.BytesIO(data))
    total = len(reader.pages)
    # invoice i spans pages starts[i] .. starts[i+1]-1 (last one runs to the end)
    spans = [(s, (starts[i + 1] - 1) if i + 1 < len(starts) else total)
             for i, s in enumerate(starts)]

    out = []
    for i in range(0, len(spans), max_invoices):
        batch = spans[i:i + max_invoices]
        first, last = batch[0][0], batch[-1][1]
        writer = PdfWriter()
        for p in range(first, last + 1):
            writer.add_page(reader.pages[p - 1])   # pypdf is 0-based
        buf = io.BytesIO()
        writer.write(buf)
        out.append((buf.getvalue(), first))
        log.info("  chunk %d: pages %d-%d (%d invoice(s))", len(out), first, last, len(batch))
    return out


def extract_invoice(client: genai.Client, filename: str, data: bytes,
                    mime: str = "application/pdf",
                    extra_prompt: str = "", temperature: float | None = None,
                    max_attempts: int = 2) -> tuple[list[dict], dict]:
    # ponytail: even at temperature=0 Gemini occasionally emits broken JSON on very
    # large documents (server-side variance), so retry the whole call a couple of
    # times. Every attempt's tokens are billed, so all of them count toward usage.
    #
    # max_attempts=1 is passed for a batch holding several invoices: re-sending the same
    # request reproduces the same bad JSON, and we have a better recovery for it — cut the
    # batch smaller and retry (retry_failed_by_splitting). Retrying first just burns a full
    # call. A single invoice can't be split, so it keeps both attempts.
    usage = {"input": 0, "output": 0}
    last_exc: Exception | None = None
    for attempt in range(1, max_attempts + 1):
        log.info("  Gemini call start: %s (%.0f KB, attempt %d)", filename, len(data) / 1024, attempt)
        t0 = time.perf_counter()
        response = client.models.generate_content(
            model="gemini-2.5-flash",
            contents=[
                types.Part.from_bytes(data=data, mime_type=mime),
                EXTRACTION_PROMPT + extra_prompt,
            ],
            # temperature=0: deterministic, transcription-mode output (no creative guessing).
            # On retry use 0.2 — a bad generation at temp 0 tends to repeat verbatim
            # (server-side caching), so the retry must be allowed to differ.
            # max_output_tokens=65536: model max, so huge invoices (100+ line items) don't truncate mid-JSON.
            config=types.GenerateContentConfig(
                response_mime_type="application/json",
                temperature=(0 if attempt == 1 else 0.2) if temperature is None else temperature,
                max_output_tokens=65536,
            ),
        )
        log.info("  Gemini call done:  %s in %.1fs", filename, time.perf_counter() - t0)
        u = response.usage_metadata
        if u:
            usage["input"] += u.prompt_token_count or 0
            usage["output"] += (u.candidates_token_count or 0) + (u.thoughts_token_count or 0)
        try:
            parsed = json.loads(response.text)
            break
        except (json.JSONDecodeError, TypeError) as exc:
            last_exc = ValueError(f"Gemini returned non-JSON output: {exc}")
            log.warning("  non-JSON output for %s on attempt %d: %s", filename, attempt, exc)
    else:
        # carry the tokens burned by the failed attempts so job cost stays accurate
        last_exc.usage = usage
        raise last_exc
    parsed = _fix_mojibake(parsed)

    # ponytail: Gemini usually returns one object but occasionally wraps it in an
    # array (or a single PDF genuinely holds several invoices). Normalize to a list.
    if isinstance(parsed, dict):
        invoices = [parsed]
    elif isinstance(parsed, list):
        invoices = [item for item in parsed if isinstance(item, dict)]
    else:
        exc = ValueError(f"Gemini returned unexpected JSON shape: {type(parsed).__name__}")
        exc.usage = usage
        raise exc
    if not invoices:
        exc = ValueError("Gemini returned no invoice objects")
        exc.usage = usage
        raise exc

    for inv in invoices:
        inv["source_file"] = filename
    return invoices, usage


FIXED_COLUMNS = [
    "Source File", "Invoice No", "Invoice Date", "Seller", "Seller GSTIN", "Buyer", "Buyer GSTIN",
    "Item Description", "HSN/SAC", "Quantity", "Unit", "Rate", "Amount",
    "Taxable Value", "CGST Rate", "CGST Amount", "SGST Rate", "SGST Amount",
    "IGST Rate", "IGST Amount", "Round Off", "Grand Total", "Grand Total (Words)", "Validation",
]


def build_workbook(invoices: list[dict], errors: list[dict] | None = None) -> Workbook:
    # Everything on one sheet: one row per line item, invoice-level fields repeated
    # per row. Any document-specific "other_fields" become extra columns (union of
    # all labels seen). Failed files get a row with the Error column filled.
    errors = errors or []

    other_labels = []  # preserve first-seen order across invoices
    for inv in invoices:
        for label in (inv.get("other_fields") or {}):
            if label not in other_labels:
                other_labels.append(label)

    header = FIXED_COLUMNS + other_labels + (["Error"] if errors else [])

    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"
    ws.append(header)

    def invoice_base(inv):
        return {
            "Source File": inv.get("source_file"), "Invoice No": inv.get("invoice_no"),
            "Invoice Date": inv.get("invoice_date"), "Seller": inv.get("seller_name"),
            "Seller GSTIN": inv.get("seller_gstin"), "Buyer": inv.get("buyer_name"),
            "Buyer GSTIN": inv.get("buyer_gstin"), "Taxable Value": inv.get("taxable_value"),
            "CGST Rate": inv.get("cgst_rate"), "CGST Amount": inv.get("cgst_amount"),
            "SGST Rate": inv.get("sgst_rate"), "SGST Amount": inv.get("sgst_amount"),
            "IGST Rate": inv.get("igst_rate"), "IGST Amount": inv.get("igst_amount"),
            "Round Off": inv.get("round_off"), "Grand Total": inv.get("grand_total"),
            "Grand Total (Words)": inv.get("grand_total_words"),
            "Validation": _validate_invoice(inv),
            **{label: (inv.get("other_fields") or {}).get(label) for label in other_labels},
        }

    for inv in invoices:
        base = invoice_base(inv)
        items = inv.get("items") or [{}]  # emit at least one row even if no line items
        for it in items:
            row = dict(base)
            row.update({
                "Item Description": it.get("description"), "HSN/SAC": it.get("hsn"),
                "Quantity": it.get("qty"), "Unit": it.get("unit"),
                "Rate": it.get("rate"), "Amount": it.get("amount"),
            })
            ws.append([_cell(row.get(col)) for col in header])

    for e in errors:
        row = {"Source File": e.get("source_file"), "Error": e.get("error")}
        ws.append([_cell(row.get(col)) for col in header])

    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    for col_cells in ws.columns:
        lengths = [len(str(c.value)) if c.value is not None else 0 for c in col_cells]
        max_len = max(lengths, default=10)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    return wb


def _is_rate_limit(exc: Exception) -> bool:
    # ponytail: string-match the SDK error; covers 429 / RESOURCE_EXHAUSTED / quota wording.
    msg = str(exc).lower()
    return any(s in msg for s in ("429", "resource_exhausted", "rate limit", "ratelimit", "quota", "too many requests"))


async def _process_chunk(client, chunk, concurrency, total, attempt, on_file_progress=None,
                         attempts_by_idx=None, custom_fields_prompt=""):
    """Run a chunk of (idx, filename, data) through Gemini, at most `concurrency` at once.

    `attempts_by_idx` overrides how many times a given work item may be re-sent on bad
    JSON. A batch of several invoices gets 1 — splitting it is the better recovery.

    Returns (invoices, hard_errors, rate_limited). Rate-limited files are returned
    separately so the caller can retry them as a smaller sub-chunk.
    """
    sem = asyncio.Semaphore(concurrency)
    attempts_by_idx = attempts_by_idx or {}

    async def worker(i, name, data, mime):
        file_start = time.perf_counter()
        try:
            async with sem:
                if on_file_progress:
                    on_file_progress(i, "processing")
                result, usage = await asyncio.to_thread(
                    extract_invoice, client, name, data, mime, custom_fields_prompt, None,
                    attempts_by_idx.get(i, 2))
            log.info("[%d/%d] OK %s (%d rows) in %.1fs", i, total, name, len(result),
                     time.perf_counter() - file_start)
            # Don't report "done" here — the reconciliation re-read and the final DB
            # save still have to happen. Files.status should only ever become "done"
            # from save_invoices_to_db(), once there's an actual Invoice record to show
            # for it; a premature "done" here defeats resume_interrupted_jobs() (which
            # only looks for processing/queued/retrying) if the process dies in between.
            return ("ok", i, name, data, mime, (result, usage))
        except Exception as exc:
            if _is_rate_limit(exc):
                log.warning("[%d/%d] RATE-LIMITED %s (%.1fs)", i, total, name,
                            time.perf_counter() - file_start)
                if on_file_progress:
                    on_file_progress(i, "retrying")
                return ("rate", i, name, data, mime, exc)
            log.warning("[%d/%d] FAILED %s: %s (%.1fs)", i, total, name, exc,
                        time.perf_counter() - file_start)
            if on_file_progress:
                on_file_progress(i, "failed")
            return ("fail", i, name, data, mime, exc)

    done = await asyncio.gather(*(worker(i, name, data, mime) for i, name, data, mime in chunk))

    invoices, hard_errors, rate_limited = [], [], []
    usage = {"input": 0, "output": 0}
    for status, i, name, data, mime, payload in done:
        if status == "ok":
            inv, u = payload
            # tag each invoice with the upload it came from. The filename can't be the
            # join key — two uploads in one batch may share a name.
            for one in inv:
                one["file_idx"] = i
            invoices.extend(inv)
            usage["input"] += u["input"]
            usage["output"] += u["output"]
        elif status == "rate":
            rate_limited.append((i, name, data, mime))
        else:
            hard_errors.append({"source_file": name, "error": str(payload), "file_idx": i})
            # failed extractions still consumed tokens — keep the job cost honest
            u = getattr(payload, "usage", None)
            if u:
                usage["input"] += u["input"]
                usage["output"] += u["output"]
    return invoices, hard_errors, rate_limited, usage


async def process_adaptive(client, chunk, concurrency, total, attempt=1, on_file_progress=None,
                           attempts_by_idx=None, custom_fields_prompt=""):
    """Process a chunk; on rate-limit, retry the rate-limited files as a smaller
    sub-chunk (halved concurrency) after an exponential backoff, down to sequential."""
    invoices, errors, rate_limited, usage = await _process_chunk(
        client, chunk, concurrency, total, attempt, on_file_progress, attempts_by_idx,
        custom_fields_prompt)

    if rate_limited:
        if attempt > RATE_LIMIT_RETRIES:
            log.error("Giving up on %d file(s) after %d rate-limit retries", len(rate_limited), RATE_LIMIT_RETRIES)
            for i, name, _, _ in rate_limited:
                errors.append({"source_file": name, "file_idx": i,
                               "error": f"rate limit after {RATE_LIMIT_RETRIES} retries"})
        else:
            backoff = 2 ** attempt  # 2, 4, 8, 16s
            sub_conc = max(1, concurrency // 2)
            log.warning("Rate-limited %d file(s); retry as sub-chunk (concurrency %d, attempt %d) after %ds",
                        len(rate_limited), sub_conc, attempt + 1, backoff)
            await asyncio.sleep(backoff)
            sub_inv, sub_err, sub_usage = await process_adaptive(
                client, rate_limited, sub_conc, total, attempt + 1, on_file_progress,
                attempts_by_idx, custom_fields_prompt)
            invoices.extend(sub_inv)
            errors.extend(sub_err)
            usage["input"] += sub_usage["input"]
            usage["output"] += sub_usage["output"]
    return invoices, errors, usage


async def reread_unreconciled(client, invoices: list[dict], chunk, custom_fields_prompt="") -> tuple[list[dict], dict]:
    """Second opinion on the documents whose figures don't add up.

    A failed check is usually the document's own doing (a charge we can't model) but it
    can also be a genuine misread — most often on a phone photo. So re-read those
    documents once, telling the model exactly what didn't reconcile, and keep the new
    answer only if it reconciles better than the old one. Costs tokens only for the
    documents that failed, and never makes the result worse.
    """
    usage = {"input": 0, "output": 0}
    by_idx = {i: (name, data, mime) for i, name, data, mime in chunk}

    def bad(inv_list):
        return [v for v in (_validate_invoice(i) for i in inv_list) if v != "OK"]

    flagged = sorted({inv["file_idx"] for inv in invoices
                      if inv.get("file_idx") in by_idx and _validate_invoice(inv) != "OK"})
    if not flagged:
        return invoices, usage

    log.info("Verification pass: re-reading %d document(s) whose figures didn't reconcile", len(flagged))
    for idx in flagged:
        name, data, mime = by_idx[idx]
        mine = [inv for inv in invoices if inv.get("file_idx") == idx]
        issues = "; ".join(sorted(set(bad(mine))))
        try:
            fresh, u = await asyncio.to_thread(
                extract_invoice, client, name, data, mime,
                VERIFY_SUFFIX.format(issues=issues) + custom_fields_prompt, 0.2)
        except Exception as exc:
            log.warning("  re-read of %s failed (%s); keeping the first reading", name, exc)
            u = getattr(exc, "usage", None)
            if u:
                usage["input"] += u["input"]
                usage["output"] += u["output"]
            continue
        usage["input"] += u["input"]
        usage["output"] += u["output"]
        for one in fresh:
            one["file_idx"] = idx

        if len(bad(fresh)) < len(bad(mine)):
            log.info("  re-read of %s reconciles (%d unreconciled -> %d); using it",
                     name, len(bad(mine)), len(bad(fresh)))
            invoices = [inv for inv in invoices if inv.get("file_idx") != idx] + fresh
        else:
            log.info("  re-read of %s still doesn't reconcile; the document itself likely "
                     "doesn't add up — keeping the original reading", name)
    invoices.sort(key=lambda i: i.get("file_idx", 0))
    return invoices, usage


# In-memory job store. Extraction runs in the background so the HTTP request returns
# immediately (no nginx/proxy timeout) and the result downloads from a real URL
# (native browser download — works on iOS and in-app browsers, unlike blob URLs).
# ponytail: single-process, in-memory — jobs die on restart and aren't shared across
# workers. Fine for one uvicorn worker; swap for Redis if you scale to several.
JOBS: dict[str, dict] = {}
JOBS_MAX = 100  # cap memory; drop the oldest jobs beyond this

XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _new_job(jid: str | None = None) -> str:
    jid = jid or uuid.uuid4().hex   # resume reuses the original id, so Files docs still match
    JOBS[jid] = {"status": "processing", "xlsx": None, "error": None,
                 "ok": 0, "failed": 0, "created": time.time(), "task": None,
                 "tokens": None, "cost_usd": None, "cost_inr": None,
                 "files": []}  # [{"name": str, "status": "queued"|"processing"|"retrying"|"done"|"failed"}]
    if len(JOBS) > JOBS_MAX:
        oldest = min(JOBS, key=lambda k: JOBS[k]["created"])
        JOBS.pop(oldest, None)
    return jid


def _local_starts(starts: list[int], first: int, last: int) -> list[int]:
    """Invoice start pages of a sub-PDF, renumbered 1-based within that sub-PDF."""
    return [s - first + 1 for s in starts if first <= s <= last]


async def expand_large_pdfs(client, chunk) -> tuple[list, dict, dict, dict]:
    """Replace each oversized PDF with several sub-PDFs of whole invoices.

    A single Gemini response is capped at 65536 output tokens, so a PDF holding ~19
    invoices truncates mid-JSON and fails. We ask Gemini where each invoice begins, then
    cut the file on those boundaries — never mid-invoice, so a 2-page invoice stays whole.

    Returns (new_chunk, work_to_file, work_meta, usage). Small files pass through
    untouched, so they cost and behave exactly as before.
    """
    usage = {"input": 0, "output": 0}
    new_chunk, work_to_file, work_meta = [], {}, {}
    next_id = max((i for i, _, _, _ in chunk), default=0) + 1

    for i, name, data, mime in chunk:
        pages = pdf_page_count(data) if mime == "application/pdf" else 0
        if pages <= SPLIT_PAGES:
            new_chunk.append((i, name, data, mime))
            work_to_file[i] = i
            continue
        try:
            starts, u = await asyncio.to_thread(find_invoice_pages, client, name, data)
            usage["input"] += u["input"]
            usage["output"] += u["output"]
            parts = await asyncio.to_thread(split_pdf_by_invoice, data, starts, MAX_INVOICES_PER_CALL)
        except Exception:
            # boundary pass failed — fall back to the whole file (it may still truncate,
            # but we must not lose the document)
            log.exception("  could not split %s; extracting it whole", name)
            new_chunk.append((i, name, data, mime))
            work_to_file[i] = i
            continue

        log.info("Split %s: %d pages, %d invoice(s) -> %d call(s)",
                 name, pages, len(starts), len(parts))
        bounds = [(s, (starts[k + 1] - 1) if k + 1 < len(starts) else pages)
                  for k, s in enumerate(starts)]
        for part_data, first_page in parts:
            part_pages = pdf_page_count(part_data)
            last_page = first_page + part_pages - 1
            wid = next_id
            next_id += 1
            new_chunk.append((wid, f"{name} (p{first_page}+)", part_data, mime))
            work_to_file[wid] = i          # every piece maps back to the one library file
            # offset/pages let a page number the model reports inside this piece be
            # translated back to a page of the file the user actually uploaded
            work_meta[wid] = {"data": part_data, "mime": mime, "name": name,
                              "offset": first_page, "pages": part_pages,
                              "starts": _local_starts(starts, first_page, last_page)}
        del bounds
    return new_chunk, work_to_file, work_meta, usage


def resolve_pages(invoices, work_meta):
    """Translate each invoice's page number into a page of the file the user uploaded.

    The model is asked which page of the PDF *it was given* an invoice starts on, but that
    PDF is often one slice of a bigger upload, renumbered from 1. Without this the cards
    can't be matched to the preview: the pieces come back in whatever order they finish in,
    so page 1's invoice can end up 15th in the list.

    Falls back to the boundary pass's own start pages (which we already trust — they're what
    the file was cut on) when the model's page numbers don't hold up. Mutates in place.
    """
    by_wid = {}
    for inv in invoices:
        by_wid.setdefault(inv.get("file_idx"), []).append(inv)

    for wid, group in by_wid.items():
        meta = work_meta.get(wid, {})
        offset = meta.get("offset", 1)          # unsplit upload: the model saw the whole file
        pages = meta.get("pages")
        starts = [offset + s - 1 for s in meta.get("starts", [])]   # back to absolute pages
        local = [inv.get("page") for inv in group]
        usable = (all(isinstance(p, int) and not isinstance(p, bool) and 1 <= p <= (pages or p)
                      for p in local)
                  and local == sorted(local))
        for k, inv in enumerate(group):
            if usable:
                inv["page"] = offset + inv["page"] - 1
            elif len(group) == len(starts):
                # both lists are in document order, so the k-th invoice starts at the k-th boundary
                inv["page"] = starts[k]
            else:
                inv.pop("page", None)           # unknown beats wrong — the card just shows no page


async def retry_failed_by_splitting(client, errors, work_to_file, work_meta, depth=1, custom_fields_prompt=""):
    """A chunk that returned unparseable JSON gets cut smaller and retried.

    Truncation isn't the only way a call fails — the model sometimes emits malformed JSON
    for a particular document, and repeating the same request just reproduces it. Cutting
    the batch down isolates the awkward invoice instead of losing the whole batch with it.

    Returns (recovered_invoices, still_failed, usage).
    """
    usage = {"input": 0, "output": 0}
    splittable = [e for e in errors
                  if len(work_meta.get(e.get("file_idx"), {}).get("starts", [])) > 1]
    if not splittable or depth > 3:
        return [], errors, usage

    sub_chunk = []
    next_id = max(work_to_file, default=0) + 1
    for e in splittable:
        wid = e["file_idx"]
        meta = work_meta[wid]
        starts = meta["starts"]
        half = max(1, len(starts) // 2)
        parts = await asyncio.to_thread(split_pdf_by_invoice, meta["data"], starts, half)
        log.warning("Retrying %s as %d smaller piece(s) (%d invoice(s) did not parse)",
                    meta["name"], len(parts), len(starts))
        for part_data, first in parts:
            part_pages = pdf_page_count(part_data)
            last = first + part_pages - 1
            nid = next_id
            next_id += 1
            sub_chunk.append((nid, f"{meta['name']} (retry p{first}+)", part_data, meta["mime"]))
            work_to_file[nid] = work_to_file[wid]      # same library file
            # this piece is cut from a piece: its pages are numbered from 1 again, so the
            # offsets compose back to the original upload
            work_meta[nid] = {"data": part_data, "mime": meta["mime"], "name": meta["name"],
                              "offset": meta.get("offset", 1) + first - 1, "pages": part_pages,
                              "starts": _local_starts(starts, first, last)}

    # a piece still holding several invoices can be cut again, so it doesn't need a retry
    # either. A single-invoice piece keeps both attempts — splitting can no longer help it.
    sub_attempts = {wid: 1 for wid, m in work_meta.items()
                    if wid in {w for w, _, _, _ in sub_chunk} and len(m["starts"]) > 1}

    invoices, errs, u = await process_adaptive(client, sub_chunk, MAX_CONCURRENCY, len(sub_chunk),
                                               attempts_by_idx=sub_attempts,
                                               custom_fields_prompt=custom_fields_prompt)
    usage["input"] += u["input"]
    usage["output"] += u["output"]

    # anything still broken gets cut smaller again, until it's a single invoice
    deeper_inv, deeper_err, u2 = await retry_failed_by_splitting(
        client, errs, work_to_file, work_meta, depth + 1, custom_fields_prompt)
    usage["input"] += u2["input"]
    usage["output"] += u2["output"]

    still_failed = [e for e in errors if e not in splittable] + deeper_err
    return invoices + deeper_inv, still_failed, usage


async def _run_job(jid: str, client, chunk, pre_errors, total, stored=None, custom_fields_prompt=""):
    batch_start = time.perf_counter()
    try:
        by_name = {s["filename"]: s.get("_id") for s in (stored or [])}
        names_by_idx = {s["idx"]: s["filename"] for s in (stored or [])}

        # big PDFs become several sub-PDFs, each a separate Gemini call
        chunk, work_to_file, work_meta, split_usage = await expand_large_pdfs(client, chunk)
        # how many pieces each file was cut into — it isn't "done" until all land
        pending = {}
        for wid, parent in work_to_file.items():
            pending[parent] = pending.get(parent, 0) + 1

        def file_progress(i, status):
            j = JOBS.get(jid)
            parent = work_to_file.get(i, i)
            if j is None or not (0 <= parent - 1 < len(j["files"])):
                return
            if status == "done":
                pending[parent] = pending.get(parent, 1) - 1
                if pending[parent] > 0:
                    return          # other pieces of this file are still running
            j["files"][parent - 1]["status"] = status
            # mirror into Mongo so the library shows live progress to a user who
            # closed the tab and came back
            fid = by_name.get(j["files"][parent - 1]["name"])
            if fid is not None:
                asyncio.create_task(asyncio.to_thread(set_file_status, fid, status))

        # a batch of several invoices gets one shot: if its JSON won't parse, re-sending it
        # reproduces the same garbage — splitting it smaller is the recovery that works
        attempts_by_idx = {wid: 1 for wid, m in work_meta.items() if len(m["starts"]) > 1}

        invoices, extract_errors, usage = await process_adaptive(
            client, chunk, MAX_CONCURRENCY, len(chunk), on_file_progress=file_progress,
            attempts_by_idx=attempts_by_idx, custom_fields_prompt=custom_fields_prompt)
        usage["input"] += split_usage["input"]
        usage["output"] += split_usage["output"]

        # a batch whose JSON wouldn't parse gets cut smaller and retried, so one awkward
        # invoice doesn't take the whole batch down with it
        recovered, extract_errors, retry_usage = await retry_failed_by_splitting(
            client, extract_errors, work_to_file, work_meta, custom_fields_prompt=custom_fields_prompt)
        invoices.extend(recovered)
        usage["input"] += retry_usage["input"]
        usage["output"] += retry_usage["output"]

        # second opinion on anything whose figures don't add up (billed, so fold in the tokens)
        invoices, verify_usage = await reread_unreconciled(client, invoices, chunk, custom_fields_prompt)
        usage["input"] += verify_usage["input"]
        usage["output"] += verify_usage["output"]

        # page numbers are still local to whichever slice the model saw — fix them before
        # file_idx is remapped, since that's the key into work_meta
        resolve_pages(invoices, work_meta)

        # map the sub-PDFs' results back onto the file they came from, and restore the
        # real filename (the pieces carry a "(p7+)" suffix)
        for rec in invoices + extract_errors:
            wid = rec.get("file_idx")
            parent = work_to_file.get(wid, wid)
            rec["file_idx"] = parent
            if parent in names_by_idx:
                rec["source_file"] = names_by_idx[parent]
        # page order, so the cards read down the document the way the user sees it
        invoices.sort(key=lambda inv: (inv.get("file_idx", 0), inv.get("page") or 0))

        errors = pre_errors + extract_errors
        wb = build_workbook(invoices, errors)
        buf = io.BytesIO()
        wb.save(buf)
        cost_usd = (usage["input"] * GEMINI_INPUT_PER_1M + usage["output"] * GEMINI_OUTPUT_PER_1M) / 1_000_000
        cost = {"job_cost_usd": round(cost_usd, 4), "job_cost_inr": round(cost_usd * USD_TO_INR, 2),
                "job_tokens": usage}
        # persist to MongoDB (best-effort; never fails the job)
        await asyncio.to_thread(save_invoices_to_db, jid, invoices, errors, cost, stored)
        job = JOBS.get(jid)
        if job is None:  # evicted while running
            return
        job.update(status="done", xlsx=buf.getvalue(), ok=len(invoices), failed=len(errors),
                   tokens=usage, cost_usd=round(cost_usd, 4), cost_inr=round(cost_usd * USD_TO_INR, 2))
        log.info("Job %s done in %.1fs (%d ok, %d failed) — %d in + %d out tokens = $%.4f (₹%.2f)",
                 jid, time.perf_counter() - batch_start, len(invoices), len(errors),
                 usage["input"], usage["output"], cost_usd, cost_usd * USD_TO_INR)
    except Exception as exc:
        log.exception("Job %s crashed", jid)
        if jid in JOBS:
            JOBS[jid].update(status="error", error=str(exc))


def require_token(authorization: str = Header(default="")):
    token = authorization.removeprefix("Bearer ").strip()
    if token not in TOKENS:
        raise HTTPException(status_code=401, detail="invalid or missing token")


class LoginBody(BaseModel):
    email: str
    password: str


@app.post("/login")
def login(body: LoginBody):
    email_ok = hmac.compare_digest(body.email.strip().lower(), APP_USER.lower())
    password_ok = hmac.compare_digest(body.password, APP_PASSWORD)
    if email_ok and password_ok:
        token = uuid.uuid4().hex
        TOKENS.add(token)
        log.info("login OK for %s", body.email)
        return {"token": token}
    log.warning("login FAILED for %s", body.email)
    raise HTTPException(status_code=401, detail="invalid email or password")


@app.post("/extract", dependencies=[Depends(require_token)])
async def extract(files: list[UploadFile] = File(...), custom_fields: str = Form("[]")):
    api_key = os.environ.get("GEMINI_API_KEY")
    if not api_key:
        raise HTTPException(status_code=500, detail="GEMINI_API_KEY not set in environment/.env")

    try:
        custom_fields_list = json.loads(custom_fields) or []
    except (json.JSONDecodeError, TypeError):
        custom_fields_list = []
    custom_fields_prompt = build_custom_fields_prompt(custom_fields_list)

    client = genai.Client(api_key=api_key,
                          http_options=types.HttpOptions(timeout=GEMINI_TIMEOUT_MS))
    total = len(files)
    log.info("=== /extract received %d file(s), %d custom field(s) ===", total, len(custom_fields_list))

    # Read uploads into memory now (the request body is only available here), reject
    # non-PDFs up front, then hand off to a background task.
    jid = _new_job()
    chunk = []          # (idx, filename, data, mime) for valid PDFs and photos
    pre_errors = []
    stored = []         # what the library will list
    file_order = []     # (idx, filename) for every upload, in order — powers the UI's per-file list
    image_mimes = {"image/jpeg", "image/png", "image/webp"}
    scan_no = 0
    for i, f in enumerate(files, 1):
        log.info("[%d/%d] queued %s (%s)", i, total, f.filename, f.content_type)
        file_order.append((i, f.filename))
        title = f.filename
        if f.content_type == "application/pdf":
            data, mime = await f.read(), "application/pdf"
        elif f.content_type in image_mimes:
            # camera scan: auto-crop + deskew the photo before it goes to Gemini
            raw = await f.read()
            data = await asyncio.to_thread(preprocess_photo, raw)
            # preprocess re-encodes to JPEG on success; on fallback the bytes are untouched
            mime = "image/jpeg" if data is not raw else f.content_type
            # a blurry photo yields confident-looking nonsense — refuse it instead
            sharpness = await asyncio.to_thread(photo_sharpness, data)
            if sharpness < BLUR_THRESHOLD:
                log.warning("[%d/%d] REJECTED %s — too blurry (sharpness %.0f < %.0f)",
                            i, total, f.filename, sharpness, BLUR_THRESHOLD)
                pre_errors.append({"source_file": f.filename, "file_idx": i,
                                   "error": "photo is too blurry to read — retake it in better "
                                            "light, holding the phone steady"})
                continue
            # phones name photos IMG_2026…; give the user something human until the
            # invoice number is known
            scan_no += 1
            stamp = datetime.now().strftime("%-d %b %Y, %-I:%M %p")
            title = f"Scan · {stamp}" + (f" #{scan_no}" if scan_no > 1 else "")
        else:
            pre_errors.append({"source_file": f.filename, "file_idx": i,
                               "error": f"not a PDF or image (content-type: {f.content_type})"})
            continue
        chunk.append((i, f.filename, data, mime))
        # keep the bytes Gemini actually saw — that's what the library previews
        path = await asyncio.to_thread(store_file, jid, i, f.filename, data, mime)
        stored.append({"idx": i, "filename": f.filename, "title": title,
                       "mime": mime, "size": len(data), "path": path})

    if not chunk and not pre_errors:
        raise HTTPException(status_code=400, detail="No files were uploaded")

    # register the files now, so they're visible (and resumable) before extraction finishes
    stored = await asyncio.to_thread(create_file_docs, jid, stored)

    pre_error_names = {e["source_file"] for e in pre_errors}
    JOBS[jid]["files"] = [
        {"name": name, "status": "failed" if name in pre_error_names else "queued"}
        for _, name in file_order
    ]
    # keep a reference to the task so it isn't garbage-collected mid-flight
    JOBS[jid]["task"] = asyncio.create_task(
        _run_job(jid, client, chunk, pre_errors, total, stored, custom_fields_prompt))
    return {"job_id": jid}


@app.get("/status/{job_id}", dependencies=[Depends(require_token)])
def job_status(job_id: str):
    job = JOBS.get(job_id)
    if job is None:
        raise HTTPException(status_code=404, detail="unknown or expired job")
    done_files = sum(1 for f in job["files"] if f["status"] in ("done", "failed"))
    return {"status": job["status"], "error": job["error"], "ok": job["ok"], "failed": job["failed"],
            "tokens": job["tokens"], "cost_usd": job["cost_usd"], "cost_inr": job["cost_inr"],
            "done_files": done_files, "total_files": len(job["files"]), "files": job["files"]}


@app.get("/download/{job_id}")
def job_download(job_id: str):
    job = JOBS.get(job_id)
    if job is None:
        raise HTTPException(status_code=404, detail="unknown or expired job")
    if job["status"] == "processing":
        raise HTTPException(status_code=409, detail="still processing")
    if job["status"] == "error":
        raise HTTPException(status_code=500, detail=job["error"] or "extraction failed")
    return StreamingResponse(
        io.BytesIO(job["xlsx"]),
        media_type=XLSX_MEDIA,
        headers={"Content-Disposition": "attachment; filename=invoices_extracted.xlsx"},
    )


# --- Library: the stored files and their extracted data ---------------------

def _files_coll():
    coll = _mongo_collection(MONGO_FILES_COLLECTION)
    if coll is None:
        raise HTTPException(status_code=503, detail="MONGODB_URI not set — the library needs a database")
    return coll


def _oid(value: str):
    from bson import ObjectId
    from bson.errors import InvalidId
    try:
        return ObjectId(value)
    except (InvalidId, TypeError):
        raise HTTPException(status_code=404, detail="not found")


def _jsonable(doc: dict) -> dict:
    """Mongo doc -> JSON: ObjectId and datetime aren't JSON-serializable."""
    out = {}
    for k, v in doc.items():
        if isinstance(v, datetime):
            out[k] = v.isoformat()
        elif type(v).__name__ == "ObjectId":
            out[k] = str(v)
        else:
            out[k] = v
    return out


def _file_or_404(fid: str) -> dict:
    doc = _files_coll().find_one({"_id": _oid(fid)})
    if doc is None:
        raise HTTPException(status_code=404, detail="unknown file")
    return doc


def _invoices_of(fid: str) -> list[dict]:
    return list(_mongo_collection().find({"file_id": _oid(fid)}).sort("_id", 1))


class ExportBody(BaseModel):
    file_ids: list[str]


# lives under /files, and ahead of /files/{fid}, so "export" isn't read as a file id.
# The deployed nginx only proxies a fixed list of paths, and a top-level /export was not
# on it — the request never reached this process.
@app.post("/files/export", dependencies=[Depends(require_token)])
def export_selected(body: ExportBody):
    """One workbook from any number of chosen documents."""
    if not body.file_ids:
        raise HTTPException(status_code=400, detail="no files selected")
    oids = [_oid(f) for f in body.file_ids]
    rows = list(_mongo_collection().find({"file_id": {"$in": oids}}).sort("_id", 1))
    invoices = [r for r in rows if not r.get("error")]
    errors = [{"source_file": r.get("source_file"), "error": r["error"]}
              for r in rows if r.get("error")]
    if not invoices and not errors:
        raise HTTPException(status_code=404, detail="nothing extracted from the selected files")
    buf = io.BytesIO()
    build_workbook(invoices, errors).save(buf)   # already unions other_fields across invoices
    buf.seek(0)
    return StreamingResponse(buf, media_type=XLSX_MEDIA,
                             headers={"Content-Disposition": 'attachment; filename="invoices_export.xlsx"'})


# the invoice fields a search looks through — everything a person would type from memory
SEARCH_FIELDS = ("invoice_no", "seller_name", "buyer_name", "seller_gstin", "buyer_gstin")
MATCH_LABELS = {"invoice_no": "invoice no", "seller_name": "seller", "buyer_name": "buyer",
                "seller_gstin": "seller GSTIN", "buyer_gstin": "buyer GSTIN"}


def _search_rx(q: str) -> dict:
    """Substring, case-insensitive, and escaped — "795/26-27" is text, not a pattern.

    Unanchored on purpose: typing "ADA" has to find "ADA527/26-27". Nobody types a full name.
    """
    return {"$regex": re.escape(q.strip()), "$options": "i"}


def _library_filter(q: str) -> dict:
    """Documents matching a search: by their own name, or by an invoice they hold.

    Searching filenames alone would be useless here — every upload is called Invoice.pdf —
    so a match on any invoice inside a document matches the document.
    """
    if not q.strip():
        return {}
    rx = _search_rx(q)
    hits = _mongo_collection().distinct(
        "file_id", {"$or": [{f: rx} for f in SEARCH_FIELDS]})
    return {"$or": [{"filename": rx}, {"title": rx},
                    {"_id": {"$in": [h for h in hits if h]}}]}


def _attach_matches(docs: list[dict], q: str, cap: int = 5) -> None:
    """Say *why* each document matched: the invoices inside it that contain the term.

    A 19-invoice PDF matching on one seller looks identical to one matching on another —
    the row has to name the hit, and let the user jump straight to it. Mutates docs.
    """
    if not q.strip() or not docs:
        return
    rx = _search_rx(q)                 # same regex as the filter, so counts can't disagree
    rows = _mongo_collection().find(
        {"file_id": {"$in": [_oid(d["_id"]) for d in docs]},
         "$or": [{f: rx} for f in SEARCH_FIELDS]},
        {f: 1 for f in SEARCH_FIELDS} | {"page": 1, "file_id": 1})
    needle = q.strip().lower()
    by_file: dict[str, list] = {}
    for r in rows:
        # which field did it hit? "OLYMPIC" matches every invoice in a document when Olympic
        # is the buyer on all of them — a row that only shows the seller then looks wrong.
        hit = next((f for f in SEARCH_FIELDS if needle in str(r.get(f) or "").lower()), "")
        by_file.setdefault(str(r["file_id"]), []).append({
            "invoice_no": r.get("invoice_no") or "",
            "seller_name": r.get("seller_name") or "",
            "page": r.get("page"),          # absent on documents extracted before pages existed
            "matched_on": MATCH_LABELS.get(hit, ""),
            "matched_value": str(r.get(hit) or "") if hit else "",
        })
    for d in docs:
        mine = sorted(by_file.get(d["_id"], []), key=lambda m: m.get("page") or 0)
        d["matches"] = mine[:cap]           # a row is a list item, not a report
        d["match_count"] = len(mine)


@app.get("/files", dependencies=[Depends(require_token)])
def list_files(limit: int = 50, skip: int = 0, q: str = ""):
    """Library list, newest first, flagged when any invoice's totals don't add up."""
    limit = max(1, min(limit, 200))
    skip = max(0, skip)
    query = _library_filter(q)
    docs = [_jsonable(d) for d in
            _files_coll().find(query).sort("created_at", -1).skip(skip).limit(limit)]
    # one query for the whole page: which of these files have a validation mismatch?
    # (validation is re-run on every edit, so this reflects corrections immediately)
    flagged = _mongo_collection().distinct("file_id", {
        "file_id": {"$in": [_oid(d["_id"]) for d in docs]},
        "validation": {"$nin": ["OK", None, ""]},
    })
    flagged = {str(f) for f in flagged}
    for d in docs:
        d["needs_review"] = d["_id"] in flagged
    _attach_matches(docs, q)      # no-op without a search: the plain list pays nothing
    # total counts what the search matched, not the whole library, or the pager would
    # promise pages that don't exist
    return {"files": docs, "total": _files_coll().count_documents(query),
            "limit": limit, "skip": skip}


@app.get("/files/{fid}", dependencies=[Depends(require_token)])
def get_file(fid: str):
    """One file plus the invoices extracted from it — powers the side-by-side view."""
    return {"file": _jsonable(_file_or_404(fid)),
            "invoices": [_jsonable(d) for d in _invoices_of(fid)]}


@app.get("/files/{fid}/raw", dependencies=[Depends(require_token)])
def get_file_raw(fid: str):
    """Stream the original upload back for the preview pane.

    Newer uploads live in the private GCS bucket — this endpoint downloads the bytes
    server-side (the bucket itself has no public access) and streams them back, so
    viewing still requires a valid portal login via require_token, same as before.
    Older uploads (from before this change) still sit on local disk and are streamed
    exactly as before.
    """
    doc = _file_or_404(fid)
    raw_path = doc.get("path", "")
    if raw_path.startswith(GCS_PREFIX):
        object_name = raw_path[len(GCS_PREFIX):]
        blob = _gcs_bucket().blob(object_name)
        if not blob.exists():
            raise HTTPException(status_code=404, detail="file is no longer in storage")
        data = blob.download_as_bytes()
        return StreamingResponse(io.BytesIO(data), media_type=doc.get("mime") or "application/octet-stream")
    path = Path(raw_path).resolve()
    # the path comes from the DB, so treat it as untrusted: it must stay inside STORAGE_DIR
    if not path.is_relative_to(STORAGE_DIR) or not path.is_file():
        raise HTTPException(status_code=404, detail="file is no longer on disk")
    return FileResponse(path, media_type=doc.get("mime") or "application/octet-stream",
                        filename=doc.get("filename"))


@app.get("/files/{fid}/xlsx", dependencies=[Depends(require_token)])
def get_file_xlsx(fid: str):
    """Build the workbook from what's in the DB right now, so edits are always included."""
    doc = _file_or_404(fid)
    invoices = [i for i in _invoices_of(fid) if not i.get("error")]
    errors = [{"source_file": i.get("source_file"), "error": i["error"]}
              for i in _invoices_of(fid) if i.get("error")]
    buf = io.BytesIO()
    build_workbook(invoices, errors).save(buf)
    buf.seek(0)
    stem = Path(doc.get("filename") or "invoices").stem
    return StreamingResponse(buf, media_type=XLSX_MEDIA,
                             headers={"Content-Disposition": f'attachment; filename="{stem}.xlsx"'})


class RenameBody(BaseModel):
    title: str


@app.patch("/files/{fid}", dependencies=[Depends(require_token)])
def rename_file(fid: str, body: RenameBody):
    """Rename a document. `renamed` pins it, so a later re-extraction won't overwrite it."""
    _file_or_404(fid)
    title = body.title.strip()[:120]
    if not title:
        raise HTTPException(status_code=400, detail="title cannot be empty")
    _files_coll().update_one({"_id": _oid(fid)},
                             {"$set": {"title": title, "renamed": True}})
    return _jsonable(_file_or_404(fid))


@app.delete("/files/{fid}", dependencies=[Depends(require_token)])
def delete_file(fid: str):
    """Permanently remove a file: its bytes, its Files doc, and its extracted rows."""
    doc = _file_or_404(fid)
    oid = _oid(fid)
    path = Path(doc.get("path", "")).resolve()
    # the path comes from the DB — never delete anything outside the storage tree
    if path.is_file() and path.is_relative_to(STORAGE_DIR):
        path.unlink()
        try:
            path.parent.rmdir()   # tidy the job folder once its last file is gone
        except OSError:
            pass                  # still has siblings — fine
    removed = _mongo_collection().delete_many({"file_id": oid}).deleted_count
    _files_coll().delete_one({"_id": oid})
    log.info("deleted file %s (%s) and %d invoice row(s)", fid, doc.get("filename"), removed)
    return {"deleted": True, "invoices_removed": removed}


async def resume_interrupted_jobs():
    """Re-run any extraction that a server restart killed mid-flight.

    Files are written to disk before extraction starts, so the bytes survive; only the
    in-memory JOBS entry is lost. Anything still marked "processing" in a freshly booted
    process is by definition stale. Invoices are only inserted when a job completes, so
    resuming cannot duplicate rows — but it does re-spend Gemini tokens for that job.
    """
    coll = _mongo_collection(MONGO_FILES_COLLECTION)
    if coll is None:
        return
    stale = list(coll.find({"status": {"$in": ["processing", "queued", "retrying"]}}))
    if not stale:
        return
    api_key = os.environ.get("GEMINI_API_KEY")
    if not api_key:
        log.error("Cannot resume %d interrupted file(s): GEMINI_API_KEY not set", len(stale))
        return

    client = genai.Client(api_key=api_key, http_options=types.HttpOptions(timeout=GEMINI_TIMEOUT_MS))
    by_job: dict[str, list[dict]] = {}
    for d in stale:
        by_job.setdefault(d["job_id"], []).append(d)

    log.info("Resuming %d interrupted file(s) across %d job(s)", len(stale), len(by_job))
    for jid, docs in by_job.items():
        chunk, stored = [], []
        for i, d in enumerate(docs, 1):
            path = Path(d.get("path", "")).resolve()
            if not (path.is_file() and path.is_relative_to(STORAGE_DIR)):
                coll.update_one({"_id": d["_id"]},
                                {"$set": {"status": "failed", "error": "file missing on disk"}})
                log.warning("Cannot resume %s — bytes are gone", d.get("filename"))
                continue
            chunk.append((i, d["filename"], path.read_bytes(), d["mime"]))
            stored.append({"_id": d["_id"], "idx": i, "filename": d["filename"],
                           "mime": d["mime"], "size": d["size"], "path": d["path"]})
        if not chunk:
            continue
        _new_job(jid)   # reuse the original id so the Files docs still line up
        JOBS[jid]["files"] = [{"name": s["filename"], "status": "queued"} for s in stored]
        JOBS[jid]["task"] = asyncio.create_task(
            _run_job(jid, client, chunk, [], len(chunk), stored))
        log.info("Job %s resumed with %d file(s)", jid, len(chunk))


@app.on_event("startup")
async def _on_startup():
    STORAGE_DIR.mkdir(parents=True, exist_ok=True)
    try:
        await resume_interrupted_jobs()
    except Exception:
        log.exception("resume of interrupted jobs failed (server still starting normally)")


@app.get("/health")
def health():
    return {"status": "ok"}


@app.get("/", response_class=HTMLResponse)
def upload_page():
    # ponytail: self-contained page, no Swagger. Native multi-file input + fetch()
    # so we get a file list, a busy state, and an auto-download of the .xlsx.
    return """
    <!doctype html>
    <title>Invoice to Excel</title>
    <style>
      body { font-family: system-ui, sans-serif; max-width: 640px; margin: 3rem auto; padding: 0 1rem; }
      .drop { border: 2px dashed #888; border-radius: 8px; padding: 2rem; text-align: center; cursor: pointer; }
      .drop.over { border-color: #2563eb; background: #eff6ff; }
      ul { padding-left: 1.2rem; }
      button { margin-top: 1rem; padding: .6rem 1.2rem; font-size: 1rem; cursor: pointer; }
      button:disabled { opacity: .5; cursor: default; }
      #status { margin-top: 1rem; }
    </style>
    <h2>Invoice to Excel</h2>
    <p>Select or drop invoice PDFs (bulk supported), then extract to one Excel workbook.</p>

    <div class="drop" id="drop">Click to choose PDFs, or drag &amp; drop them here</div>
    <input type="file" id="files" accept="application/pdf" multiple hidden>
    <ul id="list"></ul>
    <button id="go" disabled>Extract to Excel</button>
    <div id="status"></div>

    <script>
      const input = document.getElementById('files');
      const drop = document.getElementById('drop');
      const list = document.getElementById('list');
      const go = document.getElementById('go');
      const status = document.getElementById('status');
      let picked = [];

      function setFiles(fileList) {
        picked = Array.from(fileList).filter(f => f.type === 'application/pdf' || f.name.toLowerCase().endsWith('.pdf'));
        list.innerHTML = picked.map(f => '<li>' + f.name + '</li>').join('');
        go.disabled = picked.length === 0;
        status.textContent = picked.length ? picked.length + ' file(s) ready' : '';
      }

      drop.onclick = () => input.click();
      input.onchange = () => setFiles(input.files);
      drop.ondragover = e => { e.preventDefault(); drop.classList.add('over'); };
      drop.ondragleave = () => drop.classList.remove('over');
      drop.ondrop = e => { e.preventDefault(); drop.classList.remove('over'); setFiles(e.dataTransfer.files); };

      const sleep = ms => new Promise(r => setTimeout(r, ms));

      go.onclick = async () => {
        go.disabled = true;
        status.textContent = 'Uploading ' + picked.length + ' file(s)...';
        const fd = new FormData();
        picked.forEach(f => fd.append('files', f));
        try {
          const res = await fetch('/extract', { method: 'POST', body: fd });
          if (!res.ok) { status.textContent = 'Error: ' + await res.text(); go.disabled = false; return; }
          const { job_id } = await res.json();
          status.textContent = 'Extracting... this can take a while for many files.';
          // poll until the background job finishes
          let s;
          while (true) {
            await sleep(2000);
            s = await (await fetch('/status/' + job_id)).json();
            if (s.status === 'done') break;
            if (s.status === 'error') { status.textContent = 'Error: ' + s.error; go.disabled = false; return; }
          }
          // native download from a real URL (works on iOS / in-app browsers)
          window.location = '/download/' + job_id;
          status.textContent = 'Done. Excel downloading.' +
            (s.cost_usd != null ? ' Gemini cost: $' + s.cost_usd + ' (Rs ' + s.cost_inr + ')' : '');
          go.disabled = false;
        } catch (err) {
          status.textContent = 'Error: ' + err;
          go.disabled = false;
        }
      };
    </script>
    """


def _self_check():
    # ponytail: smallest check that the rate-limit classifier + sub-chunk splitting hold.
    assert _is_rate_limit(Exception("429 RESOURCE_EXHAUSTED"))
    assert _is_rate_limit(Exception("Rate limit exceeded, quota"))
    assert not _is_rate_limit(ValueError("not a PDF"))
    assert not _is_rate_limit(Exception("Gemini returned non-JSON output"))

    # validation: consistent invoice passes, wrong grand total is flagged
    good = {"items": [{"amount": 100}, {"amount": 50}], "taxable_value": 150,
            "cgst_amount": 13.5, "sgst_amount": 13.5, "round_off": 0, "grand_total": 177}
    assert _validate_invoice(good) == "OK", _validate_invoice(good)
    bad = dict(good, grand_total=999)
    assert "grand total" in _validate_invoice(bad), _validate_invoice(bad)
    assert "15 characters" in _validate_invoice({"seller_gstin": "ABC"})

    # charges that have no column in our schema live in other_fields — folding them in
    # is what stops perfectly-transcribed invoices being flagged as "needs review"
    freighted = dict(good, grand_total=677, other_fields={"Freight Charges": 500})
    assert _validate_invoice(freighted) == "OK", _validate_invoice(freighted)
    discounted = dict(good, grand_total=157, other_fields={"Discount": 20})
    assert _validate_invoice(discounted) == "OK", _validate_invoice(discounted)
    # a label we don't recognise must NOT be silently absorbed
    assert _validate_invoice(dict(good, grand_total=677,
                                  other_fields={"Vehicle No": "GJ01AB1234"})) != "OK"
    # non-numeric values in other_fields must not blow up the sum
    assert _validate_invoice(dict(good, other_fields={"Freight": "n/a"})) == "OK"

    # document titles come from the invoice, not the phone's IMG_1234.jpg
    assert invoice_title([{"invoice_no": "795/26-27", "seller_name": "Olympic Decor LLP"}]) \
        == "795/26-27 — Olympic Decor LLP"
    # named after the first invoice only — the count lives on the row, not in the name
    assert invoice_title([{"invoice_no": "1"}, {"invoice_no": "2"}]) == "1"
    assert invoice_title([{}]) == ""

    # _cell flattens the list/dict values that crashed the workbook before
    assert _cell(["a", "b"]) == "a ; b"
    assert _cell({"k": "v"}) == "k: v"
    assert _cell(42) == 42 and _cell(None) is None

    async def _demo():
        calls = {"n": 0}
        seen_attempts = {}

        # fake extract that rate-limits the first 3 attempts, then succeeds
        async def fake(client, chunk, concurrency, total, attempt, on_file_progress=None,
                       attempts_by_idx=None):
            inv, err, rl = [], [], []
            for i, name, data, mime in chunk:
                calls["n"] += 1
                # record the retry budget each item was given (default 2)
                seen_attempts[i] = (attempts_by_idx or {}).get(i, 2)
                if attempt < 3:
                    rl.append((i, name, data, mime))
                else:
                    inv.append({"source_file": name})
            return inv, err, rl, {"input": 100, "output": 50}

        import types as _t
        global _process_chunk
        orig = _process_chunk
        _process_chunk = fake
        try:
            files = [(1, "a.pdf", b"", "application/pdf"), (2, "b.pdf", b"", "application/pdf")]
            inv, err, usage = await process_adaptive(None, files, 4, 2)
            assert len(inv) == 2 and not err, (inv, err)
            # 3 attempts ran (1 initial + 2 retries), each reporting 100 in / 50 out
            assert usage == {"input": 300, "output": 150}, usage
            # no budget given -> everything keeps its 2 attempts (photos, small PDFs)
            assert seen_attempts == {1: 2, 2: 2}, seen_attempts

            # a multi-invoice batch gets 1 attempt (split instead of retry); a single-invoice
            # one keeps 2, since splitting can't help it. The budget must survive the
            # rate-limit sub-chunk retry, or the skipped attempt quietly comes back.
            seen_attempts.clear()
            await process_adaptive(None, files, 4, 2, attempts_by_idx={1: 1})
            assert seen_attempts == {1: 1, 2: 2}, seen_attempts
        finally:
            _process_chunk = orig

        # page numbers: local to the slice the model saw -> pages of the uploaded file
        meta = {7: {"offset": 6, "pages": 5, "starts": [1, 3]},      # pages 6-10 of the upload
                8: {"offset": 1, "pages": 5, "starts": [1, 2]}}
        inv = [{"file_idx": 7, "page": 1}, {"file_idx": 7, "page": 3}]
        resolve_pages(inv, meta)
        assert [i["page"] for i in inv] == [6, 8], inv          # offset applied

        # model page numbers unusable -> fall back to the boundaries the file was cut on
        inv = [{"file_idx": 7, "page": 99}, {"file_idx": 7, "page": 0}]
        resolve_pages(inv, meta)
        assert [i["page"] for i in inv] == [6, 8], inv

        # unusable and no boundaries to fall back on -> no page, rather than a wrong one
        inv = [{"file_idx": 7, "page": None}, {"file_idx": 7}, {"file_idx": 7, "page": 2}]
        resolve_pages(inv, meta)
        assert all("page" not in i for i in inv), inv

        # an unsplit upload has no work_meta at all: the model's page is already correct
        inv = [{"file_idx": 3, "page": 2}]
        resolve_pages(inv, meta)
        assert inv[0]["page"] == 2, inv

        print("self-check passed")

    asyncio.run(_demo())


if __name__ == "__main__":
    import sys

    # `python main.py selfcheck` runs the asserts; `python main.py` starts the server
    # (so a production process manager that launches this file actually serves).
    if len(sys.argv) > 1 and sys.argv[1] == "selfcheck":
        _self_check()
    else:
        import uvicorn

        uvicorn.run(app, host="0.0.0.0", port=int(os.environ.get("PORT", "8000")))
